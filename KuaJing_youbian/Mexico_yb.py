import requests
import time
from lxml import etree
import openpyxl as op
from fake_useragent import UserAgent
from urllib.parse import urlencode
from KuaJing_youbian.k_j_settings import mexico_province_list


class Mexico_YouBian(object):
    def __init__(self):
        self.wb_1 = op.Workbook()  # 创建工作簿对象
        self.wb_2 = op.Workbook()  # 创建工作簿对象
        self.ws_1 = self.wb_1['Sheet']  # 创建子表
        self.ws_1.append(['邮编号', '地名/城市', '县/郡', '州/市', '纬度', '经度'])  # 添加表头
        self.ws_2 = self.wb_2['Sheet']  # 创建子表
        self.ws_2.append(['邮编号', '地名/城市', '乡/村/社区', '县/郡', '州/市', '纬度', '经度'])  # 添加表头

    def request_get(self, url):
        ua = UserAgent()
        headers = {'User-Agent': ua.random}
        request = requests.get(url=url, headers=headers)
        return request

    def xpath_html(self, response_data, xpath_str):
        response_html = etree.HTML(response_data.text)
        city_list = response_html.xpath(xpath_str)
        return city_list

    def json_data(self):
        pass

    def op_toExcel(self, data, fileName):  # openpyxl库储存数据到excel
        # wb = op.Workbook()  # 创建工作簿对象
        # ws = wb['Sheet']  # 创建子表
        # ws.append(['邮编号', '地名/城市', '县/郡', '州/市', '纬度', '经度'])  # 添加表头
        d = (data[0], data[1], data[2], data[3], data[4], data[5])
        self.ws_1.append(d)  # 每次写入一行
        self.wb_1.save(fileName)

    def op_toExce2(self, data, fileName):  # openpyxl库储存数据到excel
        # wb = op.Workbook()  # 创建工作簿对象
        # ws = wb['Sheet']  # 创建子表
        # ws.append(['邮编号', '地名/城市', '县/郡', '州/市', '纬度', '经度'])  # 添加表头
        d = (data[0], data[1], data[2], data[3], data[4], data[5], data[6])
        self.ws_2.append(d)  # 每次写入一行
        self.wb_2.save(fileName)

    def city_data(self, america_province, response_data):
        city_xpath_str = "//div[@class='col-xs-12']/div/a/text()"
        city_list = self.xpath_html(response_data, city_xpath_str)
        for city in city_list:
            city_url = 'https://www.nowmsg.com/findzip/city.asp?country=MX&state={}&county={}'.format(america_province,
                                                                                                      city)

            city_response_data = self.request_get(city_url)
            self.county_data(city_response_data, america_province, city, city_xpath_str)

    def county_data(self, city_response_data, america_province, city, city_xpath_str):
        county_list = self.xpath_html(city_response_data, city_xpath_str)
        for county in county_list:
            num = 0
            try:
                self.data_list(america_province, city, county, num)
            except Exception as e:
                num += 1
                if num < 6:
                    self.data_list(america_province, city, county, num)
                else:
                    print(num)
                    print(e)

    def data_list(self, america_province, city, county, num):
        county_xpath_str = "//table[@class='table table-hover']/tbody/tr/td/text()"
        county_url = 'https://www.nowmsg.com/findzip/postal_code.asp?country=MX&state={}&county={}&city={}'.format(
            america_province, city, county)
        county_response_data = self.request_get(county_url)
        county_data = self.xpath_html(county_response_data, county_xpath_str)
        if len(county_data) == 8:
            self.op_toExce2(county_data, 'mexico_yb_2.xlsx')
            print(america_province, city, county)
        elif len(county_data) >= 7:
            self.op_toExcel(county_data, 'mexico_yb_1.xlsx')
            print(america_province, city, county)
        else:
            if num < 6:
                num += 1
                print(num)
                time.sleep(5)
                self.data_list(america_province, city, county, num)
            else:
                print('county_data长度' + str(len(county_data)))


    def run_spider(self):
        for america_province in mexico_province_list:
            url = 'https://www.nowmsg.com/findzip/county.asp?country=MX&state={}'.format(america_province)
            response_data = self.request_get(url)
            self.city_data(america_province, response_data)


def run_MX_YouBian_spider():
    run_mc_yb = Mexico_YouBian()
    run_mc_yb.run_spider()


if __name__ == '__main__':
    run_MX_YouBian_spider()

