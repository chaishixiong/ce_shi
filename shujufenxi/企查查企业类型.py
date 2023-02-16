


from requests_html import HTMLSession

from openpyxl import load_workbook
from urllib.parse import quote
import os, xlwt, xlrd,time,random,re
from xlutils.copy import copy

session = HTMLSession()




class QCCSpider(object):
    USER_AGENT_LIST = ['Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; Hot Lingo 2.0)',
                       'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3451.0 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:57.0) Gecko/20100101 Firefox/57.0',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.71 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.2999.0 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.70 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10.4; en-US; rv:1.9.2.2) Gecko/20100316 Firefox/3.6.2',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.155 Safari/537.36 OPR/31.0.1889.174',
                       'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.1.4322; MS-RTC LM 8; InfoPath.2; Tablet PC 2.0)',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/48.0.2564.116 Safari/537.36 TheWorld 7',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36 OPR/55.0.2994.61',
                       'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; MATP; InfoPath.2; .NET4.0C; CIBA; Maxthon 2.0)',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.814.0 Safari/535.1',
                       'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; ja-jp) AppleWebKit/418.9.1 (KHTML, like Gecko) Safari/419.3',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36',
                       'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0; Touch; MASMJS)',
                       'Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.21 (KHTML, like Gecko) Chrome/19.0.1041.0 Safari/535.21',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
                       'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; Hot Lingo 2.0)',
                       'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3451.0 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:57.0) Gecko/20100101 Firefox/57.0',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.71 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.2999.0 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.70 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10.4; en-US; rv:1.9.2.2) Gecko/20100316 Firefox/3.6.2',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.155 Safari/537.36 OPR/31.0.1889.174',
                       'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.1.4322; MS-RTC LM 8; InfoPath.2; Tablet PC 2.0)',
                       'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/48.0.2564.116 Safari/537.36 TheWorld 7',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36 OPR/55.0.2994.61',
                       'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; MATP; InfoPath.2; .NET4.0C; CIBA; Maxthon 2.0)',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.814.0 Safari/535.1',
                       'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; ja-jp) AppleWebKit/418.9.1 (KHTML, like Gecko) Safari/419.3',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36',
                       'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0; Touch; MASMJS)',
                       'Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.21 (KHTML, like Gecko) Chrome/19.0.1041.0 Safari/535.21',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4093.3 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko; compatible; Swurl) Chrome/77.0.3865.120 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
                       'Mozilla/5.0 (X11; Linux x86_64; rv:68.0) Gecko/20100101 Goanna/4.7 Firefox/68.0 PaleMoon/28.16.0',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4086.0 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:75.0) Gecko/20100101 Firefox/75.0',
                       'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) coc_coc_browser/91.0.146 Chrome/85.0.4183.146 Safari/537.36',
                       'Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/537.36 (KHTML, like Gecko) Safari/537.36 VivoBrowser/8.4.72.0 Chrome/62.0.3202.84',
                       'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36 Edg/87.0.664.60',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.16; rv:83.0) Gecko/20100101 Firefox/83.0',
                       'Mozilla/5.0 (X11; CrOS x86_64 13505.63.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:68.0) Gecko/20100101 Firefox/68.0',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Safari/537.36',
                       'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
                       'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36 OPR/72.0.3815.400',
                       'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Safari/537.36',
                       ]

    def __init__(self):
        pass

    def start_url_response(self):

        wb = load_workbook('公司名.xlsx', data_only=True)

        sheet = wb['Sheet1']
        x_list = []  # 标题
        for row in sheet.iter_rows(min_row=3, min_col=2, ):
            for cell in row:

                # print(cell.value, end=' ')
                if cell.value != None:
                    x_list.append(cell.value)
                else:
                    break
                break


        for x in x_list:

            time.sleep(random.randint(1,2))
            start_url = f'https://www.qcc.com/web/search?key={quote(x)}'
            headers = {
                'cookie': 'UM_distinctid=17a0fc9f67a1c-0a1138a0a0e8bd-2363163-384000-17a0fc9f67bfff; zg_did=%7B%22did%22%3A%20%2217a0fc9f6c430-0d78d2f6aaebd6-2363163-384000-17a0fc9f6c5ce4%22%7D; _uab_collina=162376253226282941731779; QCCSESSID=nfraeucal7td4n63cm4kqre2b1; acw_tc=6f06b41d16238281894435910e4ebbeb5e82345db67187f5b55aa56298; CNZZDATA1254842228=1819538414-1623758195-%7C1623826973; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201623828191702%2C%22updated%22%3A%201623829647050%2C%22info%22%3A%201623762532040%2C%22superProperty%22%3A%20%22%7B%5C%22%E5%BA%94%E7%94%A8%E5%90%8D%E7%A7%B0%5C%22%3A%20%5C%22%E4%BC%81%E6%9F%A5%E6%9F%A5%E7%BD%91%E7%AB%99%5C%22%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%5C%22%24utm_source%5C%22%3A%20%5C%22baidu1%5C%22%2C%5C%22%24utm_medium%5C%22%3A%20%5C%22cpc%5C%22%2C%5C%22%24utm_term%5C%22%3A%20%5C%22pzsy%5C%22%7D%22%2C%22referrerDomain%22%3A%20%22www.baidu.com%22%2C%22zs%22%3A%200%2C%22sc%22%3A%200%2C%22cuid%22%3A%20%22f695bd8c3bc3b4018ba6e1eb75edbaa0%22%7D',
                'referer': f'https://www.qcc.com/web/search?key={quote(x)}',
                'user-agent': random.choice(self.USER_AGENT_LIST)
            }
            response = session.get(start_url, headers=headers)
            # print(response.content.decode())
            try:
                next_url = response.html.xpath('//div[@class="maininfo"]/a/@href')[0]
            except:
                continue
            # print(next_url)

            self.next_url_response(next_url,x)
            # break

    def next_url_response(self, next_url,x):
        itme = {}
        start_next_url = next_url + '#base'  # 基本信息页面
        #https://www.qcc.com/firm/bd9d1df350a643af0747fc2795e47248.html#base
        #https://www.qcc.com/firm/bd9d1df350a643af0747fc2795e47248.html#base
        # print(start_next_url)
        headers = {
            # 'cookie': 'UM_distinctid=17c2f814f2f1f4-0004e481eace5c-b7a193d-1a76a2-17c2f814f30a97; acw_tc=6f301f1816365456550052938e80e1c6ab39f39209a366089180c45a53; QCCSESSID=tliphq1n3u5psnrnljeq1htul5; CNZZDATA1254842228=1145463347-1632881362-https%253A%252F%252Fwww.baidu.com%252F%7C1636543775; zg_did=%7B%22did%22%3A%20%2217c2f814ec3c45-0bb3dc59a1ec1d-b7a193d-1a76a2-17c2f814ec4e1d%22%7D; zg_294c2ba1ecc244809c552f8f6fd2a440=%7B%22sid%22%3A%201636545656192%2C%22updated%22%3A%201636545726046%2C%22info%22%3A%201636541488391%2C%22superProperty%22%3A%20%22%7B%5C%22%E5%BA%94%E7%94%A8%E5%90%8D%E7%A7%B0%5C%22%3A%20%5C%22%E4%BC%81%E6%9F%A5%E6%9F%A5%E7%BD%91%E7%AB%99%5C%22%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%5C%22%24utm_source%5C%22%3A%20%5C%22baidu1%5C%22%2C%5C%22%24utm_medium%5C%22%3A%20%5C%22cpc%5C%22%2C%5C%22%24utm_term%5C%22%3A%20%5C%22pzsy%5C%22%7D%22%2C%22referrerDomain%22%3A%20%22www.baidu.com%22%2C%22cuid%22%3A%20%22d8888f592af3a3b44873f55ae9b6c677%22%2C%22zs%22%3A%200%2C%22sc%22%3A%200%2C%22firstScreen%22%3A%201636545656192%7D',
            'cookie': 'UM_distinctid=17a0fc9f67a1c-0a1138a0a0e8bd-2363163-384000-17a0fc9f67bfff; zg_did=%7B%22did%22%3A%20%2217a0fc9f6c430-0d78d2f6aaebd6-2363163-384000-17a0fc9f6c5ce4%22%7D; _uab_collina=162376253226282941731779; QCCSESSID=nfraeucal7td4n63cm4kqre2b1; acw_tc=6f06b41d16238281894435910e4ebbeb5e82345db67187f5b55aa56298; CNZZDATA1254842228=1819538414-1623758195-%7C1623826973; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201623828191702%2C%22updated%22%3A%201623829647050%2C%22info%22%3A%201623762532040%2C%22superProperty%22%3A%20%22%7B%5C%22%E5%BA%94%E7%94%A8%E5%90%8D%E7%A7%B0%5C%22%3A%20%5C%22%E4%BC%81%E6%9F%A5%E6%9F%A5%E7%BD%91%E7%AB%99%5C%22%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%5C%22%24utm_source%5C%22%3A%20%5C%22baidu1%5C%22%2C%5C%22%24utm_medium%5C%22%3A%20%5C%22cpc%5C%22%2C%5C%22%24utm_term%5C%22%3A%20%5C%22pzsy%5C%22%7D%22%2C%22referrerDomain%22%3A%20%22www.baidu.com%22%2C%22zs%22%3A%200%2C%22sc%22%3A%200%2C%22cuid%22%3A%20%22f695bd8c3bc3b4018ba6e1eb75edbaa0%22%7D',
            'referer': f'https://www.qcc.com/web/search?key={quote(x)}',
            'user-agent': random.choice(self.USER_AGENT_LIST)
        }

        response = session.get(start_next_url,headers=headers).html

        a = ''.join(response.xpath('//*[@id="cominfo"]/div[2]/table/tr[9]/td[2]/div/span[1]/a[1]/text()'))

        # print(a)

        itme['公司名称'] = x
        itme['公司地址'] = a

        self.save_excel(itme)

    def save_excel(self,itme):

        data_list = [itme['公司名称'],itme['公司地址']]

        data = {
            '基本详情': data_list
        }
        os_path_1 = os.getcwd() + '/数据/'
        if not os.path.exists(os_path_1):
            os.mkdir(os_path_1)
        # os_path = os_path_1 + self.os_path_name + '.xls'

        os_path = os_path_1 + '数据.xls'

        if not os.path.exists(os_path):
            # 创建新的workbook（其实就是创建新的excel）
            workbook = xlwt.Workbook(encoding='utf-8')
            # 创建新的sheet表
            worksheet1 = workbook.add_sheet("基本详情", cell_overwrite_ok=True)
            borders = xlwt.Borders()  # Create Borders
            """定义边框实线"""
            borders.left = xlwt.Borders.THIN
            borders.right = xlwt.Borders.THIN
            borders.top = xlwt.Borders.THIN
            borders.bottom = xlwt.Borders.THIN
            borders.left_colour = 0x40
            borders.right_colour = 0x40
            borders.top_colour = 0x40
            borders.bottom_colour = 0x40
            style = xlwt.XFStyle()  # Create Style
            style.borders = borders  # Add Borders to Style
            """居中写入设置"""
            al = xlwt.Alignment()
            al.horz = 0x02  # 水平居中
            al.vert = 0x01  # 垂直居中
            style.alignment = al
            # 合并 第0行到第0列 的 第0列到第13列
            '''基本详情13'''
            # worksheet1.write_merge(0, 0, 0, 13, '基本详情', style)
            excel_data_1 = ('公司名称','公司地址')
            for i in range(0, len(excel_data_1)):
                worksheet1.col(i).width = 2560 * 3
                #               行，列，  内容，            样式
                worksheet1.write(0, i, excel_data_1[i], style)
            workbook.save(os_path)
        # 判断工作表是否存在
        if os.path.exists(os_path):
            # 打开工作薄
            workbook = xlrd.open_workbook(os_path)
            # 获取工作薄中所有表的个数
            sheets = workbook.sheet_names()
            for i in range(len(sheets)):
                for name in data.keys():
                    worksheet = workbook.sheet_by_name(sheets[i])
                    # 获取工作薄中所有表中的表名与数据名对比
                    if worksheet.name == name:
                        # 获取表中已存在的行数
                        rows_old = worksheet.nrows
                        # 将xlrd对象拷贝转化为xlwt对象
                        new_workbook = copy(workbook)
                        # 获取转化后的工作薄中的第i张表
                        new_worksheet = new_workbook.get_sheet(i)
                        for num in range(0, len(data[name])):
                            new_worksheet.write(rows_old, num, data[name][num])
                        new_workbook.save(os_path)

        print(f'{itme}----------保存完成-------------')




if __name__ == '__main__':
    q = QCCSpider()
    q.start_url_response()
